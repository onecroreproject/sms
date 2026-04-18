from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Q, Case, When, Value, IntegerField
from django.core.paginator import Paginator
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.core.mail import send_mail
import random
from .models import Student
from .forms import StudentForm
import datetime
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

@login_required
def dashboard(request):
    queryset = Student.objects.all()
    
    # Search functionality
    search_query = request.GET.get('search')
    if search_query:
        # Prioritize results: 
        # 1. Name matches (Score 3)
        # 2. Contact matches (Score 2)
        # 3. Other matches (Score 1)
        queryset = queryset.annotate(
            search_priority=Case(
                When(name__icontains=search_query, then=Value(3)),
                When(Q(mobile__icontains=search_query) | Q(whatsapp__icontains=search_query), then=Value(2)),
                When(Q(degree__icontains=search_query) | Q(department__icontains=search_query) | Q(passed_out_year__icontains=search_query), then=Value(1)),
                default=Value(0),
                output_field=IntegerField(),
            )
        ).filter(search_priority__gt=0).order_by('-search_priority', '-created_at')
    else:
        queryset = queryset.order_by('-created_at')

    # Filter functionality
    degree_filter = request.GET.get('degree')
    dept_filter = request.GET.get('department')
    year_filter = request.GET.get('year')

    if degree_filter:
        queryset = queryset.filter(degree__iexact=degree_filter)
    if dept_filter:
        queryset = queryset.filter(department__iexact=dept_filter)
    if year_filter:
        queryset = queryset.filter(passed_out_year=year_filter)

    # Summary Stats (Based on filtered queryset as per user request)
    total_students = queryset.count()
    today = timezone.now().date()
    today_added = queryset.filter(created_at__date=today).count()

    # Duplicates Logic (Always check all students for duplicates)
    from django.db.models import Count
    duplicate_emails = Student.objects.values('email').annotate(c=Count('email')).filter(c__gt=1).values_list('email', flat=True)
    duplicate_mobiles = Student.objects.values('mobile').annotate(c=Count('mobile')).filter(c__gt=1).values_list('mobile', flat=True)
    
    # We want to show the total number of records that are duplicates
    duplicate_records = Student.objects.filter(Q(email__in=duplicate_emails) | Q(mobile__in=duplicate_mobiles))
    duplicate_count = duplicate_records.count()

    # Missing Data Logic
    missing_data_records = Student.objects.filter(
        Q(name__isnull=True) | Q(name="") |
        Q(email__isnull=True) | Q(email="") |
        Q(mobile__isnull=True) | Q(mobile="") |
        Q(whatsapp__isnull=True) | Q(whatsapp="") |
        Q(degree__isnull=True) | Q(degree="") |
        Q(department__isnull=True) | Q(department="") |
        Q(passed_out_year__isnull=True)
    )
    missing_data_count = missing_data_records.count()

    # Pagination
    paginator = Paginator(queryset, 10) # 10 records per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Get unique values for filters
    degrees = Student.objects.values_list('degree', flat=True).distinct()
    departments = Student.objects.values_list('department', flat=True).distinct()
    years = Student.objects.values_list('passed_out_year', flat=True).distinct().order_by('-passed_out_year')

    # Add student form for the modal
    form = StudentForm()

    context = {
        'students': page_obj,
        'total_students': total_students,
        'today_added': today_added,
        'duplicate_count': duplicate_count,
        'missing_data_count': missing_data_count,
        'degrees': degrees,
        'departments': departments,
        'years': years,
        'search_query': search_query,
        'selected_degree': degree_filter,
        'selected_dept': dept_filter,
        'selected_year': year_filter,
        'form': form,
    }
    return render(request, 'students/dashboard.html', context)

def student_create(request):
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Student registered successfully!')
            return redirect('dashboard')
        else:
            # If form is invalid, re-render dashboard with errors and trigger modal
            return render_dashboard_with_errors(request, form)
    return redirect('dashboard')

def render_dashboard_with_errors(request, invalid_form):
    # This ensures the dashboard is rendered correctly but with the invalid form context
    queryset = Student.objects.all()
    # (Optional: Re-apply filters if needed, but usually creation is from clear state)
    
    total_students = Student.objects.count()
    today_added = Student.objects.filter(created_at__date=timezone.now().date()).count()
    
    paginator = Paginator(queryset, 10)
    page_obj = paginator.get_page(request.GET.get('page'))
    
    degrees = Student.objects.values_list('degree', flat=True).distinct()
    departments = Student.objects.values_list('department', flat=True).distinct()
    years = Student.objects.values_list('passed_out_year', flat=True).distinct().order_by('-passed_out_year')

    context = {
        'students': page_obj,
        'total_students': total_students,
        'today_added': today_added,
        'degrees': degrees,
        'departments': departments,
        'years': years,
        'form': invalid_form,
        'show_modal': True,
    }
    return render(request, 'students/dashboard.html', context)

from django.http import JsonResponse

def student_update(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            return JsonResponse({'status': 'success', 'message': 'Student updated successfully!'})
        else:
            errors = {field: error[0] for field, error in form.errors.items()}
            return JsonResponse({'status': 'error', 'errors': errors})
    return redirect('dashboard')

def student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        student.delete()
        messages.success(request, 'Student record deleted successfully!')
    return redirect('dashboard')

def bulk_delete(request):
    if request.method == 'POST':
        student_ids = request.POST.getlist('student_ids')
        if student_ids:
            Student.objects.filter(id__in=student_ids).delete()
            messages.success(request, f'Successfully deleted {len(student_ids)} student records.')
        else:
            messages.warning(request, 'No students were selected for deletion.')
    return redirect('dashboard')

def duplicate_list(request):
    from django.db.models import Count
    duplicate_emails = Student.objects.values('email').annotate(c=Count('email')).filter(c__gt=1).values_list('email', flat=True)
    duplicate_mobiles = Student.objects.values('mobile').annotate(c=Count('mobile')).filter(c__gt=1).values_list('mobile', flat=True)
    
    queryset = Student.objects.filter(Q(email__in=duplicate_emails) | Q(mobile__in=duplicate_mobiles)).order_by('email', 'mobile')
    
    context = {
        'students': queryset,
        'is_duplicate_view': True,
    }
    return render(request, 'students/duplicates.html', context)

def import_students(request):
    if request.method == 'GET':
        return render(request, 'students/import_students.html')
        
    try:
        if request.method == 'POST':
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.warning(request, 'No file was uploaded.')
                return redirect('dashboard')
            
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, 'Please upload a valid Excel file (.xlsx or .xls)')
                return redirect('dashboard')

            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            
            # Get header row safely
            rows_iter = sheet.iter_rows(min_row=1, max_row=1)
            try:
                header_row_cells = next(rows_iter)
                header_row = [str(cell.value).strip().lower() for cell in header_row_cells if cell.value]
            except StopIteration:
                messages.error(request, 'The Excel file appears to be empty.')
                return redirect('dashboard')
            
            def get_idx(name):
                try: return header_row.index(name.lower())
                except: return -1

            idx_name = get_idx('name')
            idx_email = get_idx('email')
            idx_mobile = get_idx('mobile')
            idx_whatsapp = get_idx('whatsapp')
            idx_degree = get_idx('degree')
            idx_dept = get_idx('dept') or get_idx('department')
            idx_year = get_idx('year') or get_idx('passed out year')

            count = 0
            errors = []
            
            for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row): continue 
                
                try:
                    name = str(row[idx_name]) if idx_name != -1 and row[idx_name] else ""
                    email = str(row[idx_email]) if idx_email != -1 and row[idx_email] else ""
                    mobile = str(row[idx_mobile]) if idx_mobile != -1 and row[idx_mobile] else ""
                    whatsapp = str(row[idx_whatsapp]) if idx_whatsapp != -1 and row[idx_whatsapp] else ""
                    degree = str(row[idx_degree]) if idx_degree != -1 and row[idx_degree] else ""
                    department = str(row[idx_dept]) if idx_dept != -1 and row[idx_dept] else ""
                    year = row[idx_year] if idx_year != -1 and row[idx_year] else 0
                    
                    if not name or not email:
                        errors.append(f"Row {index}: Missing Name or Email.")
                        continue


                    
                    Student.objects.create(
                        name=name,
                        email=email,
                        mobile=mobile,
                        whatsapp=whatsapp,
                        degree=degree,
                        department=department,
                        passed_out_year=int(year) if year else 0
                    )
                    count += 1
                except Exception as row_error:
                    errors.append(f"Row {index}: {str(row_error)}")

            if count > 0:
                messages.success(request, f'Successfully imported {count} students.')
            if errors:
                for err in errors[:5]:
                    messages.warning(request, err)
                    
    except Exception as e:
        messages.error(request, f'Critical error processing file: {str(e)}')
            
    return redirect('dashboard')

def export_students_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Students_{datetime.date.today()}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    
    # Header
    columns = ['Name', 'Email', 'Mobile', 'WhatsApp', 'Degree', 'Department', 'Passed Out Year']
    ws.append(columns)
    
    # Data
    students = Student.objects.all()
    for student in students:
        ws.append([
            student.name,
            student.email,
            student.mobile,
            student.whatsapp,
            student.degree,
            student.department,
            student.passed_out_year
        ])
        
    wb.save(response)
    return response

def export_students_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Students_{datetime.date.today()}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=landscape(letter), leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
    elements = []
    
    data = [['NAME', 'EMAIL', 'MOBILE', 'WHATSAPP', 'DEGREE', 'DEPT', 'YEAR']]
    students = Student.objects.all()
    for s in students:
        data.append([
            str(s.name or ''),
            str(s.email or ''),
            str(s.mobile or ''),
            str(s.whatsapp or ''),
            str(s.degree or ''),
            str(s.department or ''),
            str(s.passed_out_year or '')
        ])
        
    # Adjust colWidths for 7 columns to fit properly on landscape page
    table = Table(data, hAlign='LEFT', colWidths=[150, 180, 90, 90, 80, 80, 50])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D9E1F2')), # Excel Light Blue Header
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black), # Thicker black grid for both vertical & horizontal
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response

# --- Auth & OTP Views ---

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            # Generate OTP
            otp = str(random.randint(100000, 999999))
            request.session['otp'] = otp
            request.session['otp_user_id'] = user.id
            request.session['otp_expiry'] = (timezone.now() + datetime.timedelta(minutes=5)).isoformat()
            
            # Send OTP to ADMIN_EMAIL
            try:
                send_mail(
                    'Login OTP Verification',
                    f'A login attempt was made for user: {username}. The OTP is: {otp}',
                    settings.DEFAULT_FROM_EMAIL,
                    [settings.ADMIN_EMAIL],
                    fail_silently=False,
                )
                messages.info(request, f'An OTP has been sent to the administrator ({settings.ADMIN_EMAIL}). Please enter it to continue.')
                return redirect('verify_otp')
            except Exception as e:
                # If console backend is used, it won't fail normally, but for SMTP it might
                messages.error(request, f'Error sending OTP: {str(e)}')
        else:
            messages.error(request, 'Invalid username or password.')
            
    return render(request, 'students/login.html')

def verify_otp(request):
    if request.method == 'POST':
        entered_otp = request.POST.get('otp')
        session_otp = request.session.get('otp')
        user_id = request.session.get('otp_user_id')
        expiry_str = request.session.get('otp_expiry')
        
        if not session_otp or not user_id:
            messages.error(request, 'Session expired or invalid. Please login again.')
            return redirect('login')
            
        expiry = timezone.datetime.fromisoformat(expiry_str)
        if timezone.now() > expiry:
            messages.error(request, 'OTP expired. Please login again.')
            return redirect('login')
            
        if entered_otp == session_otp:
            from django.contrib.auth.models import User
            user = User.objects.get(id=user_id)
            auth_login(request, user)
            
            # Clear session
            del request.session['otp']
            del request.session['otp_user_id']
            del request.session['otp_expiry']
            
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid OTP. Please try again.')
            
    return render(request, 'students/verify_otp.html')

def logout_view(request):
    auth_logout(request)
    return redirect('login')

# --- Missing Data Views ---

@login_required
def missing_data_list(request):
    queryset = Student.objects.filter(
        Q(name__isnull=True) | Q(name="") |
        Q(email__isnull=True) | Q(email="") |
        Q(mobile__isnull=True) | Q(mobile="") |
        Q(whatsapp__isnull=True) | Q(whatsapp="") |
        Q(degree__isnull=True) | Q(degree="") |
        Q(department__isnull=True) | Q(department="") |
        Q(passed_out_year__isnull=True)
    )
    
    context = {
        'students': queryset,
        'title': 'Incomplete Records',
    }
    return render(request, 'students/missing_data.html', context)
